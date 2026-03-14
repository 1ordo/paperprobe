import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RATING_COLORS = {
    "very_good": "92D050",
    "adequate": "00B0F0",
    "doubtful": "FFC000",
    "inadequate": "FF0000",
    "na": "D9D9D9",
}


def generate_excel_export(paper, assessment, standards) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "COSMIN Assessment"

    # Header
    header_font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")
    ws["A1"] = f"COSMIN Risk of Bias Assessment: {paper.title or paper.filename}"
    ws["A1"].font = header_font

    ws["A2"] = f"Authors: {paper.authors or 'N/A'}"
    ws["A3"] = f"Year: {paper.year or 'N/A'}"
    ws["A4"] = f"Status: {assessment.status}"

    # Column headers
    row = 6
    headers = ["Standard ID", "Question", "Section", "AI Rating", "AI Confidence",
               "Reviewer 1", "Reviewer 2", "Final", "AI Reasoning"]
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border

    # Build rating lookup
    rating_map = {r.standard_id: r for r in assessment.standard_ratings}

    row = 7
    for standard in standards:
        rating = rating_map.get(standard.id)

        ws.cell(row=row, column=1, value=standard.standard_number).border = thin_border
        ws.cell(row=row, column=2, value=standard.question_text).border = thin_border
        ws.cell(row=row, column=3, value=standard.section_group or "").border = thin_border

        if rating:
            for col, val in [(4, rating.ai_rating), (6, rating.reviewer1_rating),
                             (7, rating.reviewer2_rating), (8, rating.final_rating)]:
                cell = ws.cell(row=row, column=col, value=val or "")
                cell.border = thin_border
                if val and val in RATING_COLORS:
                    cell.fill = PatternFill(start_color=RATING_COLORS[val], end_color=RATING_COLORS[val], fill_type="solid")

            ws.cell(row=row, column=5, value=f"{rating.ai_confidence:.0%}" if rating.ai_confidence else "").border = thin_border
            ws.cell(row=row, column=9, value=rating.ai_reasoning or "").border = thin_border
        else:
            for col in range(4, 10):
                ws.cell(row=row, column=col, value="").border = thin_border

        row += 1

    # Auto-width columns
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 50

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_csv_export(paper, assessment, standards) -> str:
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow(["Standard ID", "Question", "Section", "AI Rating",
                     "AI Confidence", "Reviewer 1", "Reviewer 2", "Final", "AI Reasoning"])

    rating_map = {r.standard_id: r for r in assessment.standard_ratings}

    for standard in standards:
        rating = rating_map.get(standard.id)
        writer.writerow([
            standard.standard_number,
            standard.question_text,
            standard.section_group or "",
            rating.ai_rating if rating else "",
            f"{rating.ai_confidence:.2f}" if rating and rating.ai_confidence else "",
            rating.reviewer1_rating if rating else "",
            rating.reviewer2_rating if rating else "",
            rating.final_rating if rating else "",
            rating.ai_reasoning if rating else "",
        ])

    return output.getvalue()
